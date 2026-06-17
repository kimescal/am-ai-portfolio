import logging

from typing import List, Dict, Any
import json
from datetime import datetime, date
from langchain_core.tools import tool
from pydantic import BaseModel, Field
from agents.tools.sql_query import sql_query
from openpyxl import Workbook
from datetime import date, timedelta
from chinese_calendar import is_workday

from service.qiwei.zhiyu_platform import get_api_token_cached, uploadFile, sendFile
from langchain_core.runnables import RunnableConfig
import json
import os


logger = logging.getLogger(__name__)


def generateExcel(data: List[Dict[str, Any]], years_list: List[int], region: str, manager_group_map: dict = None, is_full_data: bool = False) -> str:
    """
    根据参数生成Excel文件
    列：客群或区域,客户经理,2025年度拜访次数,2025年工作日拜访率,2026以来拜访次数,2026年工作日拜访率,上周拜访次数,变化趋势
    其中2025，2026为years_list参数影响，传入什么就是什么从小到大排序
    客群或区域从region取值并排序
    
    Args:
        data: 拜访数据列表
        years_list: 年份列表
        region: 客群或区域字符串
        manager_group_map: 客户经理到客群/区域的映射
        is_full_data: 是否为全量数据标记
    
    Returns:
        生成的Excel文件路径
    """
    try:
        # 处理区域参数
        regions = [r.strip() for r in region.split(",")]
        regions.sort()  # 区域排序
        
        # 统计数据
        # 1. 按年份统计每个客户经理的拜访次数（按客群或区域维度）
        year_visit_counts = {}  # 结构: {group_or_region: {manager: {year: count}}}
        # 2. 按客群或区域统计数据
        group_data = {}  # 结构: {group_or_region: {manager: {year: count}}}
        # 3. 统计上周拜访次数（按客群或区域维度）
        last_week_visit_counts = {}  # 结构: {group_or_region: {manager: count}}
        
        # 计算上周的开始和结束日期（最近的7天，从7天前到1天前）
        today = date.today()
        # 上周结束日期：昨天
        last_week_end = today - timedelta(days=1)
        # 上周开始日期：7天前
        last_week_start = today - timedelta(days=7)
        
        for item in data:
            # 获取必要字段
            cust_manager = item.get("cust_manager")
            data_year = item.get("data_year")
            
            if not cust_manager or data_year is None:
                continue
            
            # 处理客户经理（可能有多个）
            managers = [m.strip() for m in cust_manager.split(",") if m.strip()]
            
            for manager in managers:
                # 根据manager_group_map获取客户经理所属的客群或区域
                item_group = manager_group_map.get(manager, "未知客群/区域") if manager_group_map else "未知客群/区域"
                
                # 初始化数据结构 - 按客群或区域统计每个客户经理的拜访次数
                if item_group not in year_visit_counts:
                    year_visit_counts[item_group] = {}
                if manager not in year_visit_counts[item_group]:
                    year_visit_counts[item_group][manager] = {}
                if data_year not in year_visit_counts[item_group][manager]:
                    year_visit_counts[item_group][manager][data_year] = 0
                
                # 初始化上周拜访次数的数据结构
                if item_group not in last_week_visit_counts:
                    last_week_visit_counts[item_group] = {}
                if manager not in last_week_visit_counts[item_group]:
                    last_week_visit_counts[item_group][manager] = 0
                
                # 统计拜访次数
                year_visit_counts[item_group][manager][data_year] += 1
                
                # 统计上周拜访次数
                try:
                    # data_date时间格式为2025-06-11T00:00:00.000
                    visit_date = datetime.strptime(str(item.get("data_date")), "%Y-%m-%dT%H:%M:%S.%f").date()
                    if last_week_start <= visit_date <= last_week_end:
                        last_week_visit_counts[item_group][manager] += 1
                except ValueError:
                    # 日期格式不正确，跳过
                    pass
                
                # 按客群或区域统计
                if item_group not in group_data:
                    group_data[item_group] = {}
                if manager not in group_data[item_group]:
                    group_data[item_group][manager] = {}
                if data_year not in group_data[item_group][manager]:
                    group_data[item_group][manager][data_year] = 0
                group_data[item_group][manager][data_year] += 1
        
        # 准备Excel数据
        excel_data = []
        
        # 处理每个客群或区域
        for gr in regions:
            # 获取该客群或区域的客户经理数据
            gr_managers = group_data.get(gr, {})
            
            # 如果客群或区域没有数据，使用该客群或区域在year_visit_counts中的数据
            if not gr_managers:
                gr_managers = year_visit_counts.get(gr, {})
            
            # 处理每个客户经理
            for manager, year_counts in gr_managers.items():
                row = {
                    "客群或区域": gr,
                    "客户经理": manager
                }
                
                # 动态生成年份列
                current_year = date.today().year
                for year in years_list:
                    visit_count = year_counts.get(year, 0)
                    # 根据年份动态计算工作日数
                    if year < current_year:
                        # 往年：计算全年工作日数
                        workdays = count_cn_workdays_in_year(year)
                    else:
                        # 今年：计算从年初到今天的工作日数
                        workdays = count_cn_workdays_this_year_until_today(include_today=True)
                    visit_rate = calculate_visit_rate(visit_count, workdays)
                    
                    row[f"{year}年度拜访次数"] = visit_count
                    row[f"{year}年工作日拜访率"] = f"{visit_rate:.2f}%"
                
                # 上周拜访次数
                row["上周拜访次数"] = last_week_visit_counts.get(gr, {}).get(manager, 0)
                # 计算变化趋势：最新一年的拜访率减去次新一年的拜访率
                if len(years_list) >= 2:
                    # 获取最新年份和次新年份
                    latest_year = years_list[-1]  # 最新年份
                    second_latest_year = years_list[-2]  # 次新年份
                    
                    # 计算最新年份的拜访率
                    latest_visit_count = year_counts.get(latest_year, 0)
                    current_year = date.today().year
                    if latest_year < current_year:
                        # 往年：计算全年工作日数
                        latest_workdays = count_cn_workdays_in_year(latest_year)
                    else:
                        # 今年：计算从年初到今天的工作日数
                        latest_workdays = count_cn_workdays_this_year_until_today(include_today=True)
                    latest_visit_rate = calculate_visit_rate(latest_visit_count, latest_workdays)
                    
                    # 计算次新年份的拜访率
                    second_latest_visit_count = year_counts.get(second_latest_year, 0)
                    if second_latest_year < current_year:
                        # 往年：计算全年工作日数
                        second_latest_workdays = count_cn_workdays_in_year(second_latest_year)
                    else:
                        # 今年：计算从年初到今天的工作日数
                        second_latest_workdays = count_cn_workdays_this_year_until_today(include_today=True)
                    second_latest_visit_rate = calculate_visit_rate(second_latest_visit_count, second_latest_workdays)
                    
                    # 使用新的classify_trend函数计算趋势
                    trend = classify_trend(latest_visit_rate, second_latest_visit_rate)
                else:
                    trend = "持平"
                row["变化趋势"] = trend
                
                excel_data.append(row)
        
        # 排序：先按客群或区域，再按最新年份的拜访率从高到低排序
        if excel_data:
            # 获取最新年份
            latest_year = years_list[-1] if years_list else None
            if latest_year:
                # 按区域分组，每个区域内按最新年份的拜访率排序
                region_groups = {}
                for item in excel_data:
                    region = item["客群或区域"]
                    if region not in region_groups:
                        region_groups[region] = []
                    region_groups[region].append(item)
                
                # 对每个区域内的数据按拜访率排序
                sorted_excel_data = []
                for region in sorted(region_groups.keys()):
                    # 获取该区域的最新年份拜访率列名
                    rate_column = f"{latest_year}年工作日拜访率"
                    # 按拜访率从高到低排序
                    region_data = sorted(region_groups[region], 
                                       key=lambda x: float(x[rate_column].strip('%')) if isinstance(x[rate_column], str) else x[rate_column], 
                                       reverse=True)
                    sorted_excel_data.extend(region_data)
                excel_data = sorted_excel_data
        
        # 使用openpyxl创建Excel文件
        wb = Workbook()
        ws = wb.active
        ws.title = "拜访率分析"
        
        # 导入样式模块
        from openpyxl.styles import Alignment, Font, Border, Side
        
        # 设置居中对齐样式
        center_alignment = Alignment(horizontal='center', vertical='center')
        # 设置加粗字体样式
        bold_font = Font(bold=True)
        # 设置边框样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 生成表头
        if excel_data:
            headers = list(excel_data[0].keys())
            # 设置默认行高
            ws.row_dimensions[1].height = 25  # 表头行高
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.alignment = center_alignment  # 表头居中
                cell.font = bold_font  # 标题字体加粗
                cell.border = thin_border  # 添加边框
                # 设置列宽
                ws.column_dimensions[chr(64 + col_idx)].width = 20  # 所有列宽设为20
            
            # 填充数据
            for row_idx, row_data in enumerate(excel_data, 2):
                # 设置数据行高
                ws.row_dimensions[row_idx].height = 20
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=row_data[header])
                    cell.alignment = center_alignment  # 数据居中
                    cell.border = thin_border  # 添加边框
            
            # 确保显示网格线
            ws.sheet_view.showGridLines = True
            
            # 合并"客群或区域"列相同值的单元格
            # 获取"客群或区域"列的索引
            if "客群或区域" in headers:
                region_col_idx = headers.index("客群或区域") + 1  # 列索引从1开始
                
                # 开始合并
                start_row = 2  # 数据开始行
                current_region = None
                
                for row_idx in range(2, len(excel_data) + 2):
                    cell_value = ws.cell(row=row_idx, column=region_col_idx).value
                    
                    if current_region is None:
                        current_region = cell_value
                        start_row = row_idx
                    elif cell_value != current_region:
                        # 如果值不同，合并之前的单元格
                        if row_idx - start_row > 1:
                            merged_cell = ws.merge_cells(start_row=start_row, start_column=region_col_idx, 
                                         end_row=row_idx-1, end_column=region_col_idx)
                            # 设置合并后单元格的对齐方式和边框
                            merged_cell = ws.cell(row=start_row, column=region_col_idx)
                            merged_cell.alignment = center_alignment
                            merged_cell.border = thin_border  # 添加边框
                        # 更新当前值和起始行
                        current_region = cell_value
                        start_row = row_idx
                
                # 处理最后一组相同值
                if len(excel_data) + 1 - start_row > 1:
                    merged_cell = ws.merge_cells(start_row=start_row, start_column=region_col_idx, 
                                 end_row=len(excel_data) + 1, end_column=region_col_idx)
                    # 设置合并后单元格的对齐方式
                    merged_cell = ws.cell(row=start_row, column=region_col_idx)
                    merged_cell.alignment = center_alignment
                    merged_cell.border = thin_border  # 添加边框
        
        # 保存文件
        # 根据is_full_data标记判断是否为全量数据
        if is_full_data:
            prefix = "全量"
        else:
            # 使用区域或客群名称作为前缀，用顿号切割
            prefix = region.replace(",", "、")
        
        filename = f"{prefix}_拜访率分析_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        
        logger.info(f"Excel文件生成成功: {filename}")
        print(f"Excel文件生成成功: {filename}")
        
        return filename
    except Exception as e:
        error_message = str(e)
        logger.error(f"生成Excel文件错误: {error_message}")
        print(f"生成Excel文件错误: {error_message}")
        return None



def calculate_workdays(start_date: date, end_date: date) -> int:
    """
    计算两个日期之间的工作日数量（包含两端）。
    """
    workdays = 0
    current_date = start_date
    while current_date <= end_date:
        # 检查当前日期是否为工作日（星期一=0，星期日=6）
        if current_date.weekday() < 5:
            workdays += 1
        current_date = date(current_date.year, current_date.month, current_date.day + 1)
    return workdays


def calculate_visit_rate(visit_count: int, workdays: int) -> float:
    """
    计算拜访率。
    """
    if workdays == 0:
        return 0.0
    return round((visit_count / workdays) * 100, 2)


def classify_trend(latest_visit_rate: float, second_latest_visit_rate: float) -> str:
    """
    根据最新年份相比次年年的拜访率绝对差值对趋势进行分类。
    """
    # 计算绝对差值
    rate_change = latest_visit_rate - second_latest_visit_rate
    
    # 根据绝对差值分类
    if rate_change >= 50:
        return "大幅提升"
    elif 10 <= rate_change < 50:
        return "提升"
    else:
        # 若持平且次年年拜访率为0则为"从零启动"
        if second_latest_visit_rate == 0:
            return "从零启动"
        else:
            return "持平"


def analyze_visit_rates(data: List[Dict[str, Any]], group_or_region: str = None) -> Dict[str, Any]:
    """
    分析拜访率并对趋势进行分类。
    处理传入方法的所有数据，无时间限制。
    每条记录代表一次拜访，cust_manager字段可能包含多个经理名称（用逗号分隔）。
    """
    # 统计每个经理的拜访次数
    visit_counts = {}  # 结构: {manager: count}
    
    for item in data:

        # 获取 cust_manager 字段
        cust_manager = item.get("cust_manager")
        if not cust_manager:
            continue
        
        # 处理 cust_manager 可能包含多个经理名称的情况（逗号分隔）
        managers = [m.strip() for m in cust_manager.split(",") if m.strip()]
        
        # 每条记录代表一次拜访，每个经理都计数一次
        for manager in managers:
            if manager not in visit_counts:
                visit_counts[manager] = 0
            visit_counts[manager] += 1
    
    # 处理每个销售人员的数据
    salespersons = {}
    for name, visit_count in visit_counts.items():
        # 计算工作日数（使用与generateExcel函数一致的方法）
        # 这里使用当前年份的总工作日数作为默认值
        current_year = date.today().year
        workdays_value = count_cn_workdays_in_year(current_year)
        
        # 计算拜访率
        visit_rate = calculate_visit_rate(visit_count, workdays_value)
        
        # 分类趋势：由于只分析单个年份的数据，根据拜访率的绝对值进行分类
        # 这里假设次年年拜访率为0，使用classify_trend函数的逻辑
        trend = classify_trend(visit_rate, 0)
        
        # 存储结果
        salespersons[name] = {
            "visit_count": visit_count,
            "workdays": workdays_value,
            "visit_rate": visit_rate,
            "trend": trend
        }
    
    # 按趋势分类
    classified = {
        "大幅提升": [],
        "提升": [],
        "从零启动": [],
        "持平": []
    }
    
    for name, info in salespersons.items():
        classified[info["trend"]].append((name, info["visit_rate"]))
    
    # 按拜访率降序排序
    for trend in classified:
        classified[trend].sort(key=lambda x: x[1], reverse=True)
    
    return {
        "salespersons": salespersons,
        "classified": classified
    }


def generate_summary(analysis_result: Dict[str, Any], group_or_region: str) -> str:
    """
    根据分析结果生成摘要文本。
    """
    classified = analysis_result["classified"]
    
    # 计算总销售人员数量
    total_salespersons = sum(len(items) for items in classified.values())
    
    # 为每种趋势生成摘要
    summaries = []
    
    if classified["大幅提升"]:
        names_rates = [f"{name}（{rate:.2f}%）" for name, rate in classified["大幅提升"]]
        summaries.append(f"26年拜访率大幅提升的{len(classified['大幅提升'])}人，分别是{('、'.join(names_rates))}")
    
    if classified["提升"]:
        names_rates = [f"{name}（{rate:.2f}%）" for name, rate in classified["提升"]]
        summaries.append(f"提升的有{len(classified['提升'])}人，分别是{('、'.join(names_rates))}")
    
    if classified["从零启动"]:
        names_rates = [f"{name}（{rate:.2f}%）" for name, rate in classified["从零启动"]]
        summaries.append(f"从零启动的有{len(classified['从零启动'])}人，分别是{('、'.join(names_rates))}")
    
    if classified["持平"]:
        names_rates = [f"{name}（{rate:.2f}%）" for name, rate in classified["持平"]]
        summaries.append(f"持平的有{len(classified['持平'])}人，分别是{('、'.join(names_rates))}")
    
    # 组合摘要
    summary = f"{group_or_region}有客户经理{total_salespersons}人，{('；'.join(summaries))}。"
    
    return summary



# [ {
#   "data_date" : "2025-06-02T00:00:00.000",
#   "cust_name" : "浦银理财",
#   "dept_name" : "浦银理财权益投资部",
#   "first_strategy" : "股票策略",
#   "second_strategy" : null,
#   "thir_strategy" : null,
#   "visit_content" : "1.浦银理财今年计划重启委外业务，目前正在梳理形成内部制度。\n2.权益投资部近期关注市场其他理财子权益委外的业务的开展模式，有意愿学习并复制。\n3.权益部对FOF策略，红利股票策略以及定增策略较为感兴趣。\n4.转债策略主要由固收部以及多资产牵头推动，权益部聚焦股票策略，研究部同步在形成信用库。",
#   "extra_ctcontent" : "前期已汇报FOF策略，后续计划请投资经理路演红利股票策略以及定增策略。在浦银委外制度生效后，推动策略上会。持续跟踪浦银委外制度落地情况，定期汇报宏观策略观点以及转债策略业绩。",
#   "cust_manager" : null,
#   "create_date" : "2025-06-17T16:55:06.000",
#   "update_date" : null,
#   "creater" : null,
#   "updater" : null,
#   "plan" : null,
#   "information" : null,
#   "region" : "华东",
#   "customer_group" : null,
#   "project_id" : 1.135643824E9
# }]
@tool(return_direct=True)
async def visit_rate_analysis(group_or_region: str, date_range: str = None, config: RunnableConfig = None,) -> str:
    """
    分析拜访率并生成摘要。
    当询问拜访情况时，系统会使用 cust_manager 字段来分析客户的拜访情况。
    当询问拜访情况没有明确说时间时，不要传递 date_range 参数，系统会自动显示今年及去年的数据。
    无论查询包含多少年份，只使用最新年份的数据进行分析和返回。
    可以处理用逗号分隔的多个客群或区域。
    注意：生成的摘要话语结束必须带上"-- 详情请查看智域已推送的文件。"这句话。
    Args:
        group_or_region: 客群或区域名称，多个用逗号分隔
        date_range: 日期查询范围，格式如"2025-2026"
    Returns: 只返回最新年份的分析结果
    """
    try:
        # 设置推送的用户列表
        user_id = None
        if config and config.get("configurable"):
            user_id = config["configurable"].get("user_id")

        if not user_id:
            return "无法生成拜访情况：未获取到用户身份信息，请重新登录，或联系管理员获取权限。"

        
        # 读取配置文件，获取人员信息
        config_path = os.path.join(os.path.dirname(__file__), "dynamic_report_job_config.json")
        
        if not os.path.exists(config_path):
            return "无法生成拜访情况：配置文件不存在，请联系管理员。"
        
        with open(config_path, 'r', encoding='utf-8') as f:
            config_data = json.load(f)
        
        # 解析客群或区域
        is_full_data = False  # 标记是否为全量数据
        if not group_or_region or group_or_region == '':
            # 当参数为空时，使用硬编码的团队列表
            is_full_data = True  # 标记为全量数据
            groups_or_regions = ["银行客群", "私行客群", "企业客群", "保险客群", "华南区域", "华北区域", "华东区域", "跨境区域", "中台"]
            group_or_region = "银行客群,私行客群,企业客群,保险客群,华南区域,华北区域,华东区域,跨境区域,中台"
        else:
            groups_or_regions = [gr.strip() for gr in group_or_region.split(",")]
        
        # 收集所有相关人员和人员-客群/区域映射
        all_managers = []
        manager_group_map = {}
        for gr in groups_or_regions:
            # 检查gr是否在团队键中（字符串包含关系）
            matched_team_key = None
            for team_key in config_data.get("teams", {}):
                if gr in team_key:
                    matched_team_key = team_key
                    break
            
            if matched_team_key:
                # 客群或区域存在，获取所有成员
                team = config_data["teams"][matched_team_key]
                # 添加leader
                # if team.get("leader"):
                    # leaders = team["leader"].keys()
                    # all_managers.extend(leaders)
                    # for leader in leaders:
                    #     manager_group_map[leader] = gr
                # 添加members
                if team.get("members"):
                    members = team["members"].keys()
                    all_managers.extend(members)
                    for member in members:
                        manager_group_map[member] = gr
        
        # 去重
        all_managers = list(set(all_managers))
        
        if not all_managers:
            return "无法生成拜访情况：未找到对应客群或区域的人员信息。"
        
        # 构建SQL查询
        fields = "data_date,cust_manager"
        table = "sale_cust_dynamic_record"
        
        # 构建WHERE条件
        where_conditions = []
        
        # 添加人员条件
        if all_managers:
            # 使用REGEXP匹配逗号分隔的客户经理姓名
            # 构建正则表达式模式，将所有经理姓名组合成一个组
            managers_group = "|".join(all_managers)
            regex_pattern = f"(^|,)({managers_group})(,|$)"
            where_conditions.append(f"(cust_manager REGEXP '{regex_pattern}')")
        
        # 添加日期条件
        current_year = datetime.now().year
        last_year = current_year - 1
        
        if date_range:
            # 使用指定的日期范围
            years = date_range.split("-")
            if len(years) == 2:
                start_year = years[0]
                end_year = years[1]
                where_conditions.append(f"(data_date like '{start_year}%' or data_date like '{end_year}%')")
        else:
            # 默认使用今年和去年
            where_conditions.append(f"(data_date like '{last_year}%' or data_date like '{current_year}%')")
        
        # 构建完整SQL
        where_clause = " where " + " and ".join(where_conditions)
        sql_query_str = f"select {fields} from {table}{where_clause}"
        
        # 执行SQL查询
        sql_result = sql_query.invoke({"query": sql_query_str})

        # 解析SQL结果JSON
        try:
            result = json.loads(sql_result)
            # 检查结果是否为空
            if not result:
                return "无法生成拜访情况：未查询到相关数据或您没有权限查看该数据。"
        except Exception as e:
            return "无法生成拜访情况：数据查询失败，请联系管理员。"

        # 提取所有年份并正序存储
        years_list = await getDataYearsList(result)

        # 生成Excel文件并获取文件名
        file_path = generateExcel(result, years_list, group_or_region, manager_group_map, is_full_data)
        
        # 如果有多个客群或区域，进行分割
        groups_or_regions = [gr.strip() for gr in group_or_region.split(",")]
        
        # 找出最大的年份
        max_year = max(years_list) if years_list else None
        # 筛选出只包含最大年份数据的列表
        filtered_result = [item for item in result if item.get("data_year") == max_year]
        
        # 为每个客群或区域分析
        all_summaries = []

        for gr in groups_or_regions:
            # 分析拜访率（只使用最大年份的数据）
            analysis_result = analyze_visit_rates(filtered_result, gr)
            
            # 为当前客群或区域生成摘要
            summary = generate_summary(analysis_result, gr)
            all_summaries.append(summary)

        # 组合所有摘要
        combined_summary = " ".join(all_summaries)
        # 添加结尾语句
        combined_summary += "\n -- 详情请查看智域已推送的文件。"
        
        # 准备结果
        result_dict = {
            "result": combined_summary
        }
        result_json = json.dumps(result_dict, ensure_ascii=False)
        
        # 记录结果
        logger.info(f"拜访率分析结果: {combined_summary}")
        
        # 直接打印结果
        print(f"拜访率分析结果: {result_json}")

        # 上传并推送文件
        if file_path:
            try:
                # 获取API token
                api_token = get_api_token_cached()
                if api_token:
                    # 上传文件获取mediaId
                    media_id = await uploadFile(file_path, api_token)

                    users = [user_id]

                    # 推送文件
                    if users and media_id:
                        await sendFile(users, media_id, api_token)
                        logger.info("文件推送成功")
                else:
                    logger.warning("无API token，跳过文件推送")
            except Exception as push_error:
                logger.error(f"文件推送失败: {push_error}")

        # 返回结果为JSON
        return combined_summary
    except Exception as e:
        error_message = str(e)
        logger.error(f"拜访率分析错误: {error_message}")
        error_json = json.dumps({"error": error_message}, ensure_ascii=False)
        print(f"拜访率分析错误: {error_json}")
        return error_json


async def getDataYearsList(result):
    years_set = set()
    for item in result:
        data_date = item.get("data_date")
        if data_date:
            # 从日期字符串中提取年份 (格式: "2026-01-29")
            year = int(data_date.split("-")[0])
            years_set.add(year)
            # 添加 data_year 字段存储年份，保留原始的 data_date 字段
            item["data_year"] = year

    years_list = sorted(years_set)
    return years_list


async def ensure_sql_has_time_filter(sql_query_str: str) -> str:
    current_year = datetime.now().year
    last_year = current_year - 1

    # 检查SQL是否包含WHERE子句
    sql_lower = sql_query_str.lower()
    where_index = sql_lower.find('where')

    # 检查是否已有data_date条件
    has_data_date_condition = False
    if where_index != -1:
        # 提取WHERE子句后的部分
        where_part = sql_query_str[where_index:]
        if 'data_date' in where_part:
            has_data_date_condition = True

    # 如果没有data_date条件，添加今年和去年的时间条件
    if not has_data_date_condition:
        if where_index != -1:
            # 在现有WHERE子句后添加AND条件
            modified_sql = sql_query_str + f" and (data_date like '{last_year}%' or data_date like '{current_year}%')"
        else:
            # 添加新的WHERE子句
            modified_sql = sql_query_str + f" where (data_date like '{last_year}%' or data_date like '{current_year}%')"
        sql_query_str = modified_sql
        logger.info(f"添加时间条件后的SQL: {sql_query_str}")
    return sql_query_str

def count_cn_workdays_this_year_until_today(include_today: bool = False) -> int:
    today = date.today()
    end = today if include_today else today - timedelta(days=1)

    start = date(today.year, 1, 1)

    count = 0
    cur = start
    while cur <= end:
        if is_workday(cur):
            count += 1
        cur += timedelta(days=1)

    return count

def count_cn_workdays_in_year(year: int) -> int:
    """
    统计指定年份内的中国法定工作日总数（包含调休补班，排除法定节假日）
    :param year: 年份，例如 2023
    :return: 该年份的工作日总数
    """
    start = date(year, 1, 1)
    end = date(year, 12, 31)

    count = 0
    cur = start
    while cur <= end:
        if is_workday(cur):
            count += 1
        cur += timedelta(days=1)

    return count